import warnings
warnings.filterwarnings("ignore", category=RuntimeWarning, message=".*grpcio.*", module="opentelemetry.*")
# === FastAPI Core ===
from fastapi import FastAPI, Request, UploadFile, File, HTTPException, Depends, Form, Body, Path as FastAPIPath, Query
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

# === Security & Auth ===
from fastapi.security import OAuth2PasswordBearer
from jose import jwt, JWTError
import bcrypt
from dotenv import load_dotenv

# === Pydantic Models ===
from pydantic import BaseModel, EmailStr

# === Database & Files ===
import pymysql
import pymysql.cursors
from io import BytesIO
from openpyxl import load_workbook
from fastapi import Query

# === LLMs & Vector Store ===
import ollama
import chromadb
from llama_index.core.memory import ChatMemoryBuffer
from llama_index.llms.ollama import Ollama
from llama_index.embeddings.huggingface import HuggingFaceEmbedding
from llama_index.core import Settings, StorageContext, VectorStoreIndex, get_response_synthesizer
from llama_index.vector_stores.chroma import ChromaVectorStore
from llama_index.core.vector_stores import MetadataFilter, MetadataFilters, FilterOperator
from llama_index.core.retrievers import VectorIndexRetriever
from llama_index.core.query_engine import RetrieverQueryEngine
from llama_index.core.base.embeddings.base import BaseEmbedding
from llama_index.core.llms import ChatMessage, MessageRole

# === Utilities ===
from datetime import datetime, time, timedelta
import time as sleep_time
from urllib.parse import quote
from enum import Enum
from sentence_transformers import SentenceTransformer
from typing import Generator, Optional, List, Union
import re
import io
import os
import json
import shutil
from pathlib import Path
import itertools
import tempfile
import whisper
import openai
from datetime import datetime, timedelta
import uuid

# === REDIS ===
import redis
from fastapi import Header, Security
from typer import prompt

# === Email Stuff ===
import smtplib
import email.mime.text
import email.mime.multipart
MIMEText = email.mime.text.MIMEText
MIMEMultipart = email.mime.multipart.MIMEMultipart
import secrets
import hashlib
import base64

# === Translation ====
import argostranslate.package
import argostranslate.translate

PROJECT_ROOT = Path(__file__).resolve().parent.parent
UPLOAD_DIR = PROJECT_ROOT / "pipeline" / "data" / "raw_data"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load environment variables from .env
load_dotenv(PROJECT_ROOT / ".env")
DB_HOST = os.getenv("DB_HOST")
DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASS")
DB_NAME = os.getenv("DB_NAME")

GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
EMAIL_SENDER_USER = os.getenv("EMAIL_SENDER_USER")
EMAIL_SENDER_APP_PASSWORD = os.getenv("EMAIL_SENDER_APP_PASSWORD")
    
REMOTE_IP = os.getenv("REMOTE_IP")
LOCAL_IP = os.getenv("LOCAL_IP")

# Connect to Redis
redis_client = redis.Redis(host=LOCAL_IP, port=6380, db=0, decode_responses=True)

# Set up Ollama and HuggingFace embedding
Settings.embed_model = HuggingFaceEmbedding(model_name="intfloat/e5-large-v2")
remote_base_url = f"http://{LOCAL_IP}:11434"

# Instantiate the Ollama LLM
llm = Ollama(model="llama3.2:latest", request_timeout=120.0, temperature=0, context_window=4096, base_url=remote_base_url)
light_llm = Ollama(model="llama3.2:1b",context_window=1024,base_url=remote_base_url)
Settings.llm = llm


# Initialize ChromaDB client
PERSIST_DIR = PROJECT_ROOT / "pipeline" / "data" / "ChromaDB"
model = SentenceTransformer("intfloat/e5-large-v2")

# ChromaDB embedding function
class HFChromaEmbedding:
    def __init__(self, model):
        self.model = model

    def __call__(self, input: Union[str, List[str]]) -> Union[List[float], List[List[float]]]:
        if isinstance(input, str):
            input = [f"passage: {input}"]
            return self.model.encode(input, convert_to_numpy=True).tolist()[0]
        else:
            input = [f"passage: {text}" for text in input]
            return self.model.encode(input, convert_to_numpy=True).tolist()

    def name(self) -> str:
        return "HFChromaEmbedding-e5-large-v2"

chroma_embedding_fn = HFChromaEmbedding(model)

# LlamaIndex embedding class
class HFLlamaEmbedding(BaseEmbedding):
    model: SentenceTransformer
    def __init__(self, model):
        super().__init__(model=model)

    def _get_text_embedding(self, text: str) -> List[float]:
        return self.model.encode(f"passage: {text}", convert_to_numpy=True).tolist()

    def _get_text_embeddings(self, texts: List[str]) -> List[List[float]]:
        return self.model.encode([f"passage: {t}" for t in texts], convert_to_numpy=True).tolist()

    def _get_query_embedding(self, query: str) -> List[float]:
        return self.model.encode(f"query: {query}", convert_to_numpy=True).tolist()

    async def _aget_query_embedding(self, query: str) -> List[float]:
        return self._get_query_embedding(query)

llama_embedding_fn = HFLlamaEmbedding(model)

db = chromadb.PersistentClient(path=str(PERSIST_DIR))
chroma_collection = db.get_or_create_collection("quickstart", embedding_function=chroma_embedding_fn)
vector_store = ChromaVectorStore(chroma_collection=chroma_collection)
storage_context = StorageContext.from_defaults(vector_store=vector_store)
index = VectorStoreIndex.from_vector_store(vector_store, storage_context=storage_context, embed_model=llama_embedding_fn)
query_engine = index.as_query_engine(similarity_top_k=5, streaming=True)

# --- Utility to get user info from token/redis ---
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        session = redis_client.get(f"user_token:{user_id}")
        if not session:
            raise HTTPException(status_code=401, detail="Session expired or invalid")
        session_data = json.loads(session)
        return session_data
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    

# Prompt template
prompt_template = """
You are Verztec's AI Corporate assistant.

Your task is to extract all meaningful Corporate-related questions or concerns from the user's message.

For each item:
- Assign a concise label (e.g., "Leave entitlement", "Workplace harassment")
- Rewrite the question or concern in a short, clear form — **do not add explanations, notes, or assumptions**
- Only include relevant Corporate-related content, not exhaustive-(e.g., leave, claims, WFH, policies, pantry, etiquette, misconduct, benefits, office, ISO and QMS files or organisation-related matters)
- If the user asked about ISO files, change "ISO" to "ISO/QMS files" to include both ISO and QMS files.
- Always rewrite from the user's point of view using "I" instead of "you" unless specified
- Ignore all non-Corporate and off-topic content

⚠️ Your response must strictly follow this format:
1. **[Label]**: [Simplified question or concern]

❌ Do NOT include commentary, explanations, extra notes, or anything outside the format above.

If no Corporate-related questions or concerns are found, respond exactly: No Corporate-related questions or concerns detected.

Below is the raw user message, between <user></user> tags. Only process what's inside:

<user>
{user_input}
</user>
"""

# Pydantic model
class UserMessage(BaseModel):
    message: str
    model: str = "llama3.2:latest"  # default model
    conversation_id: Optional[str] = None
    original_message: Optional[str] = None

def get_llm(model_name: str):
    if model_name == "llama3.2:1b":
        return Ollama(model="llama3.2:1b", context_window=2048, base_url=remote_base_url)
    elif model_name == "llama3.2:latest":
        return Ollama(model="llama3.2:latest", request_timeout=120.0, context_window=4096, base_url=remote_base_url)
    elif model_name == "llama3.3":
        return Ollama(model="llama3.3", request_timeout=120.0, context_window=4096, base_url=remote_base_url)


# Call LLM and parse output
def extract_questions(user_input: str) -> list[str]:
    prompt = prompt_template.format(user_input=user_input)
    llm_model = get_llm(model_name="llama3.2:latest")  # Default to latest model
    response = llm_model.complete(prompt)
    raw_text = response.text.strip()
    print(f"[LLM] Raw response: {raw_text}")
    # If no HR-related content detected
    if "no hr-related questions or concerns are detected" in raw_text.lower():
        return []

    # Extract formatted entries
    question_list = []
    for line in raw_text.split("\n"):
        match = re.match(r"\d+\.\s+\**(.*?)\**:\s+(.*)", line.strip())
        if match:
            label, question = match.groups()
            question_list.append(f"{label.strip()}: {question.strip()}")
    print(f"[LLM] Extracted questions: {question_list}")
    return question_list

# generate autoname chat title
def generate_chat_title(user_prompt: str, model_name: str = "llama3.2:1b") -> str:
    prompt = (
        "Summarize the following user message into a short, clear chat title (max 5 words) for the following user message. "
        "Avoid generic words like 'Chat', 'Conversation', or 'Message'. Use title case, not all caps. Avoid generic titles. "
        "Do not include any extra words, headers, punctuation, or quotation marks. "
        "Output only the title itself.\n\n"
        f"{user_prompt}"
    )
    llm_model = get_llm(model_name)
    response = llm_model.complete(prompt)
    title = response.text.strip().strip('"\'').splitlines()[0]
    # Convert to title case
    return title.title()


def get_conversation_memory_from_db(conversation_id: str, user_id: int) -> ChatMemoryBuffer:
    """
    Reconstruct conversation memory from database logs.
    This is more scalable than storing in RAM.
    """
    memory = ChatMemoryBuffer.from_defaults(token_limit=1500)
    
    if not conversation_id:
        return memory
    
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            # Get recent conversation history (last 10 exchanges to limit memory usage)
            cursor.execute("""
                SELECT query, response FROM chatbot_logs 
                WHERE user_id = %s AND conversation_id = %s 
                ORDER BY created_at ASC 
                LIMIT 20
            """, (user_id, conversation_id))
            
            rows = cursor.fetchall()
            
            # Rebuild conversation memory from database
            for row in rows:
                if row['query']:
                    memory.put(ChatMessage.from_str(row['query'], role=MessageRole.USER))
                if row['response']:
                    memory.put(ChatMessage.from_str(row['response'], role=MessageRole.ASSISTANT))
                    
        return memory
    except Exception as e:
        print(f"Error loading conversation memory: {e}")
        return ChatMemoryBuffer.from_defaults(token_limit=1500)
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/process")
async def process_message(
    data: UserMessage,
    current_user: dict = Depends(get_current_user)
):
    conversation_id = data.conversation_id or str(uuid.uuid4())  # Generate a new one if not passed
    # Get conversation memory from database (not RAM)
    conversation_memory = get_conversation_memory_from_db(conversation_id, current_user["user_id"])
    
    llm_model = get_llm(model_name=data.model)
    questions = extract_questions(data.message)
    user_prompt = data.message

    if not questions:
        print("[INFO] No HR-related questions found. Generating fallback...")
        fallback_prompt = f"""You are Verztec's AI HR assistant. The user did not ask any HR-related questions or concerns.
        If the message isn't about HR (like leave, claims, benefits, or work policies), reply in a super casual, friendly manner. 
        No formal greetings, no sign-offs, no long explanations. Just a short, warm, human reply.
        If the user says something sweet or emotional (like "thank you" or "you're the best"), feel free to respond in kind—e.g., "Aww, thanks!", "You're awesome!", or use emojis.
        If you're not sure what they meant, gently ask if they have any HR-related questions, but keep it light and informal.

        Examples:
        User: Hi
        Assistant: Hey! 😊

        User: Thanks!
        Assistant: Aww, thank you! Let me know if you have any HR questions.

        User: I love you
        Assistant: Haha, you're the best! ❤️

        User: Hello
        Assistant: Hi there! How can I help you today?

        Now, reply to the user:
        """

        print(f"[Process] Querying with: {user_prompt}")
        query_engine = index.as_query_engine(streaming=True, llm=llm_model)
        chat_engine = index.as_chat_engine(
            chat_mode="context",
            memory=conversation_memory,
            system_prompt=fallback_prompt,
            query_engine=query_engine
        )
        response = chat_engine.stream_chat(f"{user_prompt}")
    
        def stream_generator():
            buffer = io.StringIO()
            try:
                for token in response.response_gen:
                    if token is not None:
                        buffer.write(token)
                        yield token
            except Exception as e:
                yield f"\n\n[ERROR streaming response: {e}]"

        # This wrapper allows us to perform logging after streaming is complete
        def streaming_with_logging():
            generator = stream_generator()
            buffer = io.StringIO()
            for token in generator:
                buffer.write(token)
                yield token
            final_response = buffer.getvalue()

            # Log to DB after full response is available
            try:
                conn = get_db()
                with conn.cursor() as cursor:
                    # sql = """
                    #     INSERT INTO chatbot_logs (user_id, conversation_id, query, response)
                    #     VALUES (%s, %s, %s, %s)
                    # """
                    # user_id = current_user["user_id"]
                    # # conversation_id = str(user_id)
                    # conversation_id = data.conversation_id or str(uuid.uuid4())
                    # cursor.execute(sql, (user_id, conversation_id, user_prompt, final_response))
                    user_id = current_user["user_id"]
                    conversation_id = data.conversation_id or str(uuid.uuid4())

                    # Check if this is the first message in the conversation
                    cursor.execute("SELECT COUNT(*) as count FROM chatbot_logs WHERE user_id = %s AND conversation_id = %s", (user_id, conversation_id))
                    row = cursor.fetchone()
                    is_first_message = row["count"] == 0

                    chat_title = None
                    if is_first_message:
                        chat_title = generate_chat_title(user_prompt)
                    else:
                        chat_title = None
                    sql = """
                        INSERT INTO chatbot_logs (user_id, username, conversation_id, query, response, title)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """
                    cursor.execute(sql, (user_id, current_user["username"], conversation_id, user_prompt, final_response, chat_title))
                conn.commit()
            finally:
                conn.close()

        return StreamingResponse(streaming_with_logging(), media_type="text/plain")
    else:
        # Stream the questions as a single JSON object (one chunk) WITH the original message
        import json
        def questions_stream():
            yield json.dumps({
                "questions": questions,
                "original_message": data.message  # Add the original message here
            })
        return StreamingResponse(questions_stream(), media_type="application/json")

    
def send_email(to_email: str, subject: str, body: str, is_html: bool = False, sender: str = GMAIL_USER):
    """Send email using Gmail SMTP"""
    try:
        # Create message
        msg = MIMEMultipart("alternative")
        msg["From"] = sender
        msg["To"] = to_email
        msg["Subject"] = subject
        
        # Add body to email
        if is_html:
            msg.attach(MIMEText(body, "html"))
        else:
            msg.attach(MIMEText(body, "plain"))
        
        # Gmail SMTP configuration
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()  # Enable security
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        
        # Send email
        text = msg.as_string()
        server.sendmail(GMAIL_USER, to_email, text)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

@app.post("/stream")
async def stream_answer(
    data: UserMessage,
    current_user: dict = Depends(get_current_user)
):
    conversation_id = data.conversation_id
    if not conversation_id:
        print("[WARNING] No conversation_id provided. Cannot log chat properly.")
    
    # Get conversation memory from database (not RAM)
    conversation_memory = get_conversation_memory_from_db(conversation_id, current_user["user_id"])
 
    # Set up the LLM based on user role
    role = current_user.get("role")
    country = current_user.get("country")
    department = current_user.get("department")

    def normalize_key(key):
        return key.strip()

    country_key = normalize_key(country)
    department_key = normalize_key(department)

    llm_model = get_llm(model_name=data.model)
    print(f"[DEBUG] User role: {role}, country: {country}, department: {department}")

    # Create the appropriate retriever based on user role
    if role == "ADMIN":
        retriever = VectorIndexRetriever(index=index, similarity_top_k=5)
        print("[DEBUG] Using UNFILTERED retriever for ADMIN")
    else:
        filters = MetadataFilters(
            filters=[
                MetadataFilter(key=country_key, value="True", operator=FilterOperator.EQ),
                MetadataFilter(key=department_key, value="True", operator=FilterOperator.EQ)
            ]
        )
        retriever = VectorIndexRetriever(index=index, similarity_top_k=5, filters=filters)
        print(f"[DEBUG] Using FILTERED retriever for {role}: {country_key}=True AND {department_key}=True")    

    test_nodes = retriever.retrieve("office plant care policy")
    print(f"[DEBUG] Plant care filter test returned {len(test_nodes)} nodes:")
    for i, node in enumerate(test_nodes):
        metadata = node.node.metadata
        country_val = metadata.get(country_key, "Unknown")
        dept_val = metadata.get(department_key, "Unknown")
        title = metadata.get('title', 'Unknown')
        print(f"  Node {i+1}: {title}")
        print(f"    {country_key}: {country_val}")
        print(f"    {department_key}: {dept_val}")
        print(f"    Should be filtered: {country_val != 'True' or dept_val != 'True'}")
    
    # Create response synthesizer and query engine with the correct retriever
    response_synthesizer = get_response_synthesizer(response_mode="tree_summarize", llm=llm_model, streaming=True)
    custom_query_engine = RetrieverQueryEngine(retriever=retriever, response_synthesizer=response_synthesizer)
    
    user_prompt = data.message
    # Use original_message for logging if available, otherwise fall back to processed message
    original_user_message = data.original_message or data.message

    system_prompt = f"""
    You are Verztec's AI HR assistant.
    You have access to the full conversation history and should use it to answer follow-up questions.
    You are able to provide information about HR policies, leave, benefits, claims, WFH, or company policies.
    Answer all questions in a friendly and human-like manner. If you are not confident about the answer,
    you should tell the user that you are not sure and suggest they contact HR directly.
    """

    print(f"[STREAM] Querying with: {data.message}")
    
    # Create chat engine with the filtered query engine
    chat_engine = index.as_chat_engine(
        chat_mode="context",
        memory=conversation_memory,
        system_prompt=system_prompt,
        query_engine=custom_query_engine
    )

    response = chat_engine.stream_chat(f"{user_prompt}")

    # Helper to check for boilerplate/empty responses
    def is_response_empty(resp):
        print("[DEBUG]: running is_empty")
        BOILERPLATE = [
            "no response",
            "no relevant information found", 
            "no answer found",
            "no content",
            "empty response",
            "there is no",
            "unfortunately",
            "i am sorry",
            "i do not",
            r"i 'm not aware",
            r"I couldn't",
        ]
        if resp is None:
            print("[DEBUG] Response is None")
            return True

        print(f"[DEBUG] Response type: {type(resp)}")

        # Check response attribute first
        if hasattr(resp, "response"):
            text = str(resp.response).strip().lower()
            print(f"[DEBUG] Found response attribute: {repr(text[:100])}")
            if text and any(phrase in text for phrase in BOILERPLATE):
                print("[DEBUG] Response marked as empty due to boilerplate in response attribute")
                return True

        # Check streaming response generator
        if hasattr(resp, "response_gen"):
            print("[DEBUG] Found response_gen attribute, attempting to peek...")
            try:
                gen = resp.response_gen
                # Peek at the first 5 tokens WITHOUT consuming them permanently
                peeked = []
                for i, token in enumerate(gen):
                    peeked.append(token)
                    if i >= 4:  # Peek at first 5 tokens (0-4)
                        break
                        
                print(f"[DEBUG] Peeked {len(peeked)} streaming chunks:")
                for i, chunk in enumerate(peeked):
                    print(f"  Chunk {i+1}: {repr(chunk)}")
                    
                joined = " ".join(str(chunk).strip().lower() for chunk in peeked if chunk and str(chunk).strip())
                print(f"[DEBUG] Joined peeked content: {repr(joined[:200])}")
                
                if not joined.strip():
                    print("[DEBUG] No content found in peeked chunks")
                    return True
                    
                for phrase in BOILERPLATE:
                    if phrase in joined:
                        print(f"[DEBUG] Found boilerplate phrase: '{phrase}' in content")
                        if len(joined) < 100:
                            return True
                            
                print("[DEBUG] Response appears to have valid content")
                
                # Store peeked tokens in a custom attribute instead of trying to replace response_gen
                resp._peeked_tokens = peeked
                resp._original_gen = gen
                return False
                
            except Exception as e:
                print(f"[DEBUG] Exception while peeking response_gen: {e}")
                import traceback
                traceback.print_exc()
                return True

        print("[DEBUG] No response_gen attribute found, assuming not empty")
        return False

    is_empty = is_response_empty(response)

    if is_empty:
        print("Empty Response")
        llm_model = get_llm(model_name="llama3.2:latest")
        fallback_text = llm_model.complete(
            f"You are Verztec's AI HR assistant. The user asked: '{user_prompt}'. "
            "You do not have information on this topic. "
            "Reply with: 'I am sorry, but I do not have information on this topic.' "
            "You can try and help the user by suggesting they contact HR directly for assistance."
        )
        raw_text = fallback_text.text.strip()

        # Add metadata to indicate this is an empty response
        response_with_metadata = f"{raw_text}\n\n__EMPTY_RESPONSE_METADATA__"

        # --- Log fallback to DB using same conversation ID ---
        try:
            conn = get_db()
            with conn.cursor() as cursor:
                sql = """
                    INSERT INTO chatbot_logs (user_id, username, conversation_id, query, response, title)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """
                user_id = current_user["user_id"]

                # Generate title for fallback response if it's the first message
                cursor.execute("SELECT COUNT(*) as count FROM chatbot_logs WHERE user_id = %s AND conversation_id = %s", (user_id, conversation_id))
                row = cursor.fetchone()
                is_first_message = row["count"] == 0
                
                chat_title = None
                if is_first_message:
                    chat_title = generate_chat_title(original_user_message)

                cursor.execute(sql, (user_id, current_user["username"], conversation_id, original_user_message, raw_text, chat_title))
            conn.commit()
        finally:
            conn.close()

        def fallback_stream():
            yield response_with_metadata
        return StreamingResponse(fallback_stream(), media_type="text/plain")

    # Check the citations for non-empty source nodes

    def simple_citation_evaluator_indices(query: str, response_text: str, source_nodes: List, llm_model) -> List[int]:
        """
        Simple approach: Ask LLM to pick the most relevant 2-3 documents and return their indices
        """
        if not source_nodes:
            return []
        
        # Ensure response_text is a string and handle edge cases
        if not isinstance(response_text, str):
            response_text = str(response_text)
        
        # Prepare document summaries
        docs_summary = []
        for i, node_score in enumerate(source_nodes[:5]):
            node = node_score.node
            doc_name = node.metadata.get("source") or f"Document_{i+1}"
            snippet = node.text[:500]  # Shorter snippet
            docs_summary.append(f"{i+1}. {doc_name}: {snippet}")
        
        prompt = f"""
            Given this user query and AI response, which documents are most relevant?

            Query: {query}
            Response: {response_text[:1000]}...

            Available documents:
            {chr(10).join(docs_summary)}

            Select the most relevant document numbers (e.g., "1,3,5"). 
            If you think none are relevant, respond with "none".
            Best number is 1 citation only.
            Only respond with the numbers, comma-separated.
            """

        try:
            llm_response = llm_model.complete(prompt)
            text = llm_response.text.strip().lower()
            if "none" in text:
                return []

            selected_nums = [int(x.strip()) for x in text.split(",") if x.strip().isdigit()]
            selected_nums = [num for num in selected_nums if 1 <= num <= len(source_nodes)]

            return selected_nums

        except Exception as e:
            print(f"[Simple Citation Index Filter] Error: {e}")
            return list(range(1, min(4, len(source_nodes) + 1)))
        
    def stream_generator():
        buffer = io.StringIO()
        response_text = ""

        # Debug: Print only document titles used
        print(f"[DEBUG] Total source nodes retrieved: {len(response.source_nodes) if hasattr(response, 'source_nodes') and response.source_nodes else 0}")
        if hasattr(response, 'source_nodes') and response.source_nodes:
            print("[DEBUG] Documents used for response generation:")
            for i, node_with_score in enumerate(response.source_nodes):
                node = node_with_score.node
                doc_name = node.metadata.get("source") or node.metadata.get("title", "Unknown")
                print(f"  {i+1}. {doc_name}")
            print()
        
        # Check if we have peeked tokens stored
        if hasattr(response, '_peeked_tokens') and hasattr(response, '_original_gen'):
            print("[DEBUG] Using restored generator with peeked tokens")
            
            # First yield the peeked tokens
            for token in response._peeked_tokens:
                if token is not None:
                    buffer.write(token)
                    response_text += token
                    yield token
            
            # Then yield the rest from the original generator
            try:
                for token in response._original_gen:
                    if token is not None:
                        buffer.write(token)
                        response_text += token
                        yield token
            except Exception as e:
                yield f"\n\n[ERROR streaming response: {e}]"
        else:
            # Fallback to original behavior if no peeked tokens
            print("[DEBUG] Using original generator")
            gen = response.response_gen
            try:
                for token in gen:
                    if token is not None:
                        buffer.write(token)
                        response_text += token
                        yield token
            except Exception as e:
                yield f"\n\n[ERROR streaming response: {e}]"

        # Signal citation processing start
        yield "\n\n__CITATION_START__"
        yield "\n"  # <-- Add this line to force a chunk between start and end

        # Rest of your citation logic...
        try:
            print("[Citation Evaluation] Evaluating citations...")
            
            selected_indices = simple_citation_evaluator_indices(
                query=user_prompt,
                response_text=str(response_text),
                source_nodes=response.source_nodes,
                llm_model=llm_model
            )

            # Signal citation processing end
            yield "__CITATION_END__"
            
        except Exception as e:
            print(f"[Citation Error] {e}")
            yield "__CITATION_END__"  # Still end citation processing even on error
            yield "\nError evaluating citations\n"
            return

        def get_real_file(filename):
            base = Path(filename).with_suffix('')
            for ext in [".pdf", ".docx", ".doc"]:
                real_path = base.with_suffix(ext)
                file_path = UPLOAD_DIR / real_path.name
                if file_path.exists():
                    return real_path.name
            return filename  # fallback

        seen_docs = []
        # Show results
        if selected_indices:
            yield "\n**📄 Most Relevant Documents:**\n"
            for idx in selected_indices:
                node = response.source_nodes[idx - 1].node
                doc_name = node.metadata.get("source") or node.metadata.get("title")
                if doc_name and doc_name not in seen_docs:
                    seen_docs.append(doc_name)
                    # Get the real file path
                    real_file = get_real_file(doc_name)
                    url = f"http://{REMOTE_IP}:8000/download/{quote(real_file)}"
                    yield f"- [{real_file}]({url})\n"
        else:
            yield "**📄 No relevant documents found**\n"


                
    # This wrapper allows us to perform logging after streaming is complete
    def streaming_with_logging():
        generator = stream_generator()
        buffer = io.StringIO()
        for token in generator:
            buffer.write(token)
            yield token
        final_response = buffer.getvalue()

        # Log to DB after full response is available
        try:
            conn = get_db()
            with conn.cursor() as cursor:
                # sql = """
                #     INSERT INTO chatbot_logs (user_id, conversation_id, query, response)
                #     VALUES (%s, %s, %s, %s)
                # """
                # user_id = current_user["user_id"]
                # # conversation_id = str(user_id)
                # # conversation_id = data.conversation_id or str(uuid.uuid4())
                # cursor.execute(sql, (user_id, conversation_id, user_prompt, final_response))
                user_id = current_user["user_id"]
                conversation_id = data.conversation_id or str(uuid.uuid4())

                # Check if this is the first message in the conversation
                cursor.execute("SELECT COUNT(*) as count FROM chatbot_logs WHERE user_id = %s AND conversation_id = %s", (user_id, conversation_id))
                row = cursor.fetchone()
                is_first_message = row["count"] == 0

                chat_title = None
                if is_first_message:
                    chat_title = generate_chat_title(original_user_message)
                else:
                    chat_title = None

                sql = """
                    INSERT INTO chatbot_logs (user_id, username, conversation_id, query, response, title)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """

                cursor.execute(sql, (user_id, current_user["username"], conversation_id, original_user_message, final_response, chat_title)) 
            conn.commit()
        finally:
            conn.close()

    return StreamingResponse(streaming_with_logging(), media_type="text/plain")

@app.post("/forward-to-hr")
async def forward_to_hr(
    data: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Forward user query to HR support via email"""
    try:
        query = data.get("query", "")
        if not query:
            raise HTTPException(status_code=400, detail="Query is required")

        # Get user info
        role = current_user.get("role")
        country = current_user.get("country")
        department = current_user.get("department")
        user_email = current_user.get("email", "unknown")

        # Send email to HR
        sender = EMAIL_SENDER_USER
        recipient = GMAIL_USER
        subject = "[HR Support Request] User Query Forwarded"
        msg_body = (
            f"A user has requested HR support for their query:\n\n"
            f"User Info:\n"
            f"- Name: {current_user.get('username', 'Unknown')}\n"
            f"- Email: {user_email}\n"
            f"- Role: {role}\n"
            f"- Country: {country}\n"
            f"- Department: {department}\n\n"
            f"User Query:\n{query}\n\n"
            f"Please follow up with the user directly at {user_email}.\n"
        )
        
        email_sent = send_email(
            to_email=recipient,
            subject=subject,
            body=msg_body,
            is_html=False,
            sender=sender
        )

        if email_sent:
            return {"message": "Your query has been forwarded to HR support. You should receive a response within 24 hours."}
        else:
            raise HTTPException(status_code=500, detail="Failed to forward query to HR")

    except Exception as e:
        print(f"[Forward to HR] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to forward query: {str(e)}")
    

# Download files
@app.get("/download/{filename}")
def download_file(filename: str):
    path = UPLOAD_DIR / filename
    return FileResponse(str(path))


@app.get("/models")
def get_models():
    # List your available models here
    return {
        "models": [
            {"name": "llama3.3", "label": "lunar ai 4"},
            {"name": "llama3.2:latest", "label": "lunar ai 3 large"},
            {"name": "llama3.2:1b", "label": "lunar ai 3 mini"},
        ]
    }



# === Auth Configuration ===
SECRET_KEY = "verztec_secret"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60


# === DB Connection ===
def get_db():
    return pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASS,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

@app.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    # Only allow ADMIN and MANAGER roles to view users
    if current_user["role"] not in ["ADMIN", "MANAGER"]:
        raise HTTPException(status_code=403, detail="Not authorized to view users")
    
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            if current_user["role"] == "MANAGER":
                # Managers only see users in their own country
                cursor.execute("""
                    SELECT user_id, username, email, role, department, country, updated_at
                    FROM users
                    WHERE country = %s
                    ORDER BY updated_at DESC
                """, (current_user["country"],))
            else:
                # Admins see all users
                cursor.execute("""
                    SELECT user_id, username, email, role, department, country, updated_at
                    FROM users
                    ORDER BY updated_at DESC
                """)
            users = cursor.fetchall()
            return users
    except Exception as e:
        print("ERROR in /users:", e)
        raise HTTPException(status_code=500, detail=f"Failed to fetch users: {str(e)}")
    finally:
        conn.close()


@app.get("/files")
async def get_files(current_user: dict = Depends(get_current_user)):
    # Only allow ADMIN and MANAGER roles to view files
    if current_user["role"] not in ["ADMIN", "MANAGER"]:
        raise HTTPException(status_code=403, detail="Not authorized to view files")
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            if current_user["role"] == "MANAGER":
                # Managers only see files for their country and department
                cursor.execute("""
                    SELECT 
                        file_id,
                        file_name,
                        uploaded_by_username AS uploaded_by,
                        countries,
                        departments,
                        upload_time
                    FROM files
                    WHERE FIND_IN_SET(%s, REPLACE(countries, ', ', ',')) > 0
                    ORDER BY upload_time DESC
                """, (current_user["country"],))
            else:
                # Admins see all files
                cursor.execute("""
                    SELECT 
                        file_id,
                        file_name,
                        uploaded_by_username AS uploaded_by,
                        countries,
                        departments,
                        upload_time
                    FROM files
                    ORDER BY upload_time DESC
                """)
            files = cursor.fetchall()
            # Convert bytes/None to string/list as needed
            for file in files:
                if isinstance(file["countries"], bytes):
                    file["countries"] = file["countries"].decode()
                if isinstance(file["departments"], bytes):
                    file["departments"] = file["departments"].decode()
            return files
    except Exception as e:
        print("ERROR in /files:", e)
        raise HTTPException(status_code=500, detail=f"Failed to fetch files: {str(e)}")
    finally:
        conn.close()

        
@app.get("/list-files", response_model=List[str])
def list_files():
    folder_path = Path(__file__).resolve().parent.parent / "pipeline" / "data" / "raw_data"
    if not folder_path.exists():
        return []
    return [f.name for f in folder_path.iterdir() if f.is_file()]

@app.delete("/delete-file/{filename}")
def delete_file(filename: str, current_user: dict = Depends(get_current_user)):
    folder_path = Path(__file__).resolve().parent.parent / "pipeline" / "data" / "raw_data"
    file_path = folder_path / filename

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    try:
        # Delete from ChromaDB and meta files first
        delete_docs_by_filename(filename)

        # --- Log deletion to file_deletion_logs and delete from DB ---
        conn = get_db()
        try:
            with conn.cursor() as cursor:
                # Fetch file info before deletion
                cursor.execute("SELECT * FROM files WHERE file_name = %s", (filename,))
                file_row = cursor.fetchone()
                if not file_row:
                    raise HTTPException(status_code=404, detail="File not found in database")

                # Log deletion
                cursor.execute("""
                    INSERT INTO file_deletion_logs (
                        deleted_file_name, deleted_file_type, department, access_level, file_path,
                        countries, batch_id, uploaded_by_username, deleted_by_username, deleted_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    file_row["file_name"],
                    file_row["file_type"],
                    file_row["department"],
                    file_row["access_level"],
                    file_row["file_path"],
                    file_row["countries"],
                    file_row.get("batch_id"),
                    # Get uploader username
                    get_username_by_user_id(file_row["uploaded_by"], cursor),
                    current_user.get("username"),
                    datetime.now()
                ))

                # Delete from files table
                cursor.execute("DELETE FROM files WHERE file_name = %s", (filename,))
            conn.commit()
        finally:
            conn.close()

        # Delete the file from disk
        file_path.unlink()
        return JSONResponse(content={"message": f"Deleted '{filename}' successfully"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/logs/{log_type}")
async def get_logs(
    log_type: str,
    current_user: dict = Depends(get_current_user)
):
    # Only ADMIN can view logs
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Not authorized to view logs")

    # Map log_type to table name and columns
    LOG_TABLES = {
        "chatbot_logs": {
            "table": "chatbot_logs",
            "columns": ["log_id", "user", "title", "query", "response", "created_at", "conversation_id"],
            "sql": """
                SELECT log_id, username AS user, title, query, response, created_at, conversation_id
                FROM chatbot_logs
                ORDER BY log_id DESC
            """
        },
        "login_logs": {
            "table": "login_logs",
            "columns": ["log_id", "user", "login_time", "status"],
            "sql": """
                SELECT log_id, username AS user, login_time, status
                FROM login_logs
                ORDER BY log_id DESC
            """
        },
        "upload_user_logs": {
            "table": "upload_user_logs",
            "columns": ["log_id", "user", "created_user", "timestamp"],
            "sql": """
                SELECT log_id, username AS user, created_username AS created_user, timestamp
                FROM upload_user_logs
                ORDER BY log_id DESC
            """
        },
        "upload_file_logs": {
            "table": "upload_file_logs",
            "columns": ["uploaded_by_username", "file_name", "department", "countries", "departments", "upload_time"],
            "sql": """
                SELECT uploaded_by_username, file_name, department, countries, departments, upload_time
                FROM upload_file_logs
                ORDER BY upload_time DESC
            """
        },
        "file_deletion_logs": {
            "table": "file_deletion_logs",
            "columns": ["log_id", "deleted_file_name", "department", "countries", "uploaded_by_username", "deleted_by_username", "deleted_at"],
            "sql": """
                SELECT log_id, deleted_file_name, department, countries, uploaded_by_username, deleted_by_username, deleted_at
                FROM file_deletion_logs
                ORDER BY log_id DESC
            """
        },
        "user_update_logs": {
            "table": "user_update_logs",
            "columns": ["log_id", "target_username", "changed_by_username", "field_name", "old_value", "new_value", "changed_at"],
            "sql": """
                SELECT log_id, target_username, changed_by_username, field_name, old_value, new_value, changed_at
                FROM user_update_logs
                ORDER BY log_id DESC
            """
        },
        "user_deletion_logs": {
            "table": "user_deletion_logs",
            "columns": ["log_id", "deleted_username", "deleted_email", "department", "role", "country", "deleted_by_username", "deleted_at"],
            "sql": """
                SELECT log_id, deleted_username, deleted_email, department, role, country, deleted_by_username, deleted_at
                FROM user_deletion_logs
                ORDER BY deleted_at DESC
            """
        },
        "file_update_logs": {
            "table": "file_update_logs",
            "columns": ["log_id", "file_id", "file_name", "changed_by_username", "field_name", "old_value", "new_value", "changed_at"],
            "sql": """
                SELECT log_id, file_id, file_name, changed_by_username, field_name, old_value, new_value, changed_at
                FROM file_update_logs
                ORDER BY changed_at DESC
            """
        },
        "password_reset_audit": {
            "table": "password_reset_audit",
            "columns": ["audit_id", "reset_type", "target_email", "target_username", "user_ip", "user_agent", "reset_token_used", "reset_at"],
            "sql": """
                SELECT audit_id, reset_type, target_email, target_username, user_ip, user_agent, reset_token_used, reset_at
                FROM password_reset_audit
                ORDER BY audit_id DESC
            """
        },
        "password_reset_tokens": {
            "table": "password_reset_tokens",
            "columns": ["user", "username", "token_hash", "expires_at", "created_at"],
            "sql": """
                SELECT username AS user, token_hash, expires_at, created_at
                FROM password_reset_tokens
                ORDER BY created_at DESC
            """
        },
    }

    if log_type not in LOG_TABLES:
        raise HTTPException(status_code=404, detail="Log type not found")

    config = LOG_TABLES[log_type]
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(config["sql"])
            rows = cursor.fetchall()
            return {"columns": config["columns"], "rows": rows}
    finally:
        conn.close()


def get_username_by_user_id(user_id, cursor):
    cursor.execute("SELECT username FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    return user["username"] if user else None

# === Models ===
class Role(str, Enum):
    ADMIN = "ADMIN"
    MANAGER = "MANAGER"
    USER = "USER"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"


# === Utility ===
def verify_password(plain_pw, hashed_pw):
    return bcrypt.checkpw(plain_pw.encode('utf-8'), hashed_pw.encode('utf-8'))

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


# === Routes ===
@app.post("/login", response_model=TokenResponse)
def login_user(user: UserLogin, request: Request):
    db = get_db()
    cursor = db.cursor(pymysql.cursors.DictCursor)
    status = "FAILURE"
    user_id = None

    try:
        cursor.execute("SELECT * FROM users WHERE email = %s", (user.email,))
        user_record = cursor.fetchone()

        if user_record and verify_password(user.password, user_record["password_hash"]):
            user_id = user_record["user_id"]
            status = "SUCCESS"

            # Store session data
            session_data = {
                "user_id": user_record["user_id"],
                "username": user_record["username"],
                "role": user_record["role"],
                "department": user_record["department"],
                "country": user_record["country"],
                "email": user_record["email"],
            }

            # Generate token
            token = create_access_token(session_data)

            # Store token in Redis with session info (as JSON string)
            redis_client.setex(
                f"user_token:{user_id}",
                ACCESS_TOKEN_EXPIRE_MINUTES * 60,
                json.dumps(session_data)
            )

            # Log successful login
            cursor.execute("""
                INSERT INTO login_logs (user_id, username, status)
                VALUES (%s, %s, %s)
                """, (user_id, user_record["username"], status))

            db.commit()
            return {"access_token": token, "token_type": "bearer"}

        else:
            # Log failed login (if user exists or not)
            if user_record:
                user_id = user_record["user_id"]
                username = user_record["username"]

            cursor.execute("""
                INSERT INTO login_logs (user_id, username, status)
                VALUES (%s, %s, %s)
                """, (user_id, username, status))

            db.commit()
            raise HTTPException(status_code=401, detail="Invalid email or password")

    finally:
        cursor.close()
        db.close()

@app.get("/profile")
def read_profile(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")

        # Check if session exists in Redis
        session = redis_client.get(f"user_token:{user_id}")
        if not session:
            raise HTTPException(status_code=401, detail="Session expired or invalid")

        return {"user_id": user_id, "session": json.loads(session)}
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


# === Conversation Logging Endpoint ===
class ChatLog(BaseModel):
    query: str
    response: str
    conversation_id: Optional[str] = None

@app.post("/log_conversation/")
async def log_conversation(
    log: ChatLog,
    current_user: dict = Depends(get_current_user)
):
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            # Check if this is the first message in the conversation
            cursor.execute("""
                SELECT COUNT(*) as count FROM chatbot_logs WHERE user_id = %s AND conversation_id = %s
            """, (current_user["user_id"], log.conversation_id))
            count_row = cursor.fetchone()
            is_first_message = (count_row["count"] == 0)

            chat_title = None
            if is_first_message and log.query:
                # Generate chat title using LLM
                chat_title = generate_chat_title(log.query)
                
            sql = """
                INSERT INTO chatbot_logs (user_id, username, conversation_id, query, response{title_col})
                VALUES (%s, %s, %s, %s, %s{title_val})
            """
            if chat_title:
                sql = sql.format(title_col=", title", title_val=", %s")
                params = (current_user["user_id"], current_user["username"], log.conversation_id or str(uuid.uuid4()), log.query, log.response, chat_title)
            else:
                sql = sql.format(title_col="", title_val="")
                params = (current_user["user_id"], current_user["username"], log.conversation_id or str(uuid.uuid4()), log.query, log.response)

            cursor.execute(sql, params)
        conn.commit()
        return {"message": "Conversation logged successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Logging failed: {str(e)}")
    finally:
        conn.close()
        
# User model
class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    department: str
    role: Role
    country: str

    
# Password hashing
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


# Upload XLSX and insert users
@app.post("/upload-xlsx")
async def upload_xlsx(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    try:
        wb = load_workbook(filename=BytesIO(await file.read()), data_only=True)
        if "users to add" in wb.sheetnames:
            sheet = wb["users to add"]
        else:
            sheet = wb.active  # fallback to first sheet if not found

        conn = get_db()
        cursor = conn.cursor()
        results = []

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            username, email, password, department, role, country = row
            line_result = {"line": i, "username": username, "status": "success", "message": "Inserted"}

            # --- Validation ---
            if not username or not role or not country or not department:
                line_result.update({"status": "error", "message": "Missing required fields"})
                results.append(line_result)
                continue

            if current_user["role"] == "MANAGER":
                if role != "USER":
                    raise HTTPException(
                        status_code=400,
                        detail=f"Managers can only create USER accounts"
                    )
                if country != current_user["country"]:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Country must be {current_user['country']}"
                    )
                if department != current_user["department"]:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Department must be {current_user['department']}"
                    )

            if not isinstance(email, str) or not (email.endswith("@verztec.com") or email.endswith("@gmail.com")): # THE @GMAIL IS ONLY FOR TESTING THE EMAIL SENDING
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid or missing email"
                )

            hashed_pw = hash_password(password)
            
            # --- Insert ---
            try:
                cursor.execute("""
                    INSERT INTO users (username, email, password_hash, department, role, country)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (username, email, hashed_pw, department, role, country))

                created_user_id = cursor.lastrowid
                cursor.execute("""
                    INSERT INTO upload_user_logs (user_id, username, created_user_id, created_username)
                    VALUES (%s, %s, %s, %s)
                    """, (current_user["user_id"], current_user["username"], created_user_id, username))

            except pymysql.err.IntegrityError:
                continue  # Skip duplicate entries

        conn.commit()
        cursor.close()
        conn.close()

        return {"message": "User data processed", "result": results}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")

def check_manager_permission(current_user, target_department, target_country):
    if current_user["role"] == "ADMIN":
        return True
    if current_user["role"] == "MANAGER":
        return (current_user["department"] == target_department and
                current_user["country"] == target_country)
    return False

@app.post("/upload-file")
async def upload_file(
    file: UploadFile = File(...),
    countries: str = Form(...),
    departments: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    countries_list = json.loads(countries)
    departments_list = json.loads(departments)

    file_path = UPLOAD_DIR / file.filename
    
    try:
        # Save the file
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Save metadata as sidecar JSON file
        meta = {
            "departments": departments_list,
            "countries": countries_list,
            "uploaded_by": current_user.get("username"),
            "upload_time": datetime.now().isoformat()
        }
        meta_path = file_path.with_suffix(file_path.suffix + ".meta.json")
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2)
        
        # Log to database
        conn = get_db()
        try:
            with conn.cursor() as cursor:
                # Get user ID
                cursor.execute("SELECT user_id FROM users WHERE username = %s", (current_user.get("username"),))
                user_result = cursor.fetchone()
                user_id = user_result["user_id"] if user_result else 1
                
                # Insert into files table
                cursor.execute("""
                    INSERT INTO files (file_name, file_type, uploaded_by, uploaded_by_username, department, access_level, file_path, 
                                    countries, departments, batch_id, upload_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    file.filename,
                    file_path.suffix.lstrip(".").lower(),
                    user_id,
                    current_user.get("username"),
                    ",".join(departments_list),
                    "FILTERED",
                    str(file_path),
                    ",".join(countries_list),
                    ",".join(departments_list),
                    None,  # batch_id is None for single upload
                    datetime.now()
                ))
                
                # INSERT INTO upload_file_logs table
                cursor.execute("""
                    INSERT INTO upload_file_logs (file_name, file_type, uploaded_by, uploaded_by_username, department, access_level, file_path, countries, departments, batch_id, upload_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    file.filename,
                    file_path.suffix.lstrip(".").lower(),
                    user_id,
                    current_user.get("username"),
                    ",".join(departments_list),
                    "FILTERED",
                    str(file_path),
                    ",".join(countries_list),
                    ",".join(departments_list),
                    None,  # batch_id is None for single upload
                    datetime.now()
                ))
                
                conn.commit()
        finally:
            conn.close()
        
        return {"message": "File uploaded successfully", "filename": file.filename}
        
    except Exception as e:
        # Clean up files if DB insert fails
        if file_path.exists():
            file_path.unlink()
        if 'meta_path' in locals() and meta_path.exists():
            meta_path.unlink()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/batch-upload-files")
async def batch_upload_files(
    files: List[UploadFile] = File(...),
    countries: str = Form(...),
    departments: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    countries_list = json.loads(countries)
    departments_list = json.loads(departments)
    
    uploaded_files = []
    failed_files = []
    
    # Create a batch metadata file to signal the watcher
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    batch_meta_path = UPLOAD_DIR / f".batch_upload_{batch_id}.processing"
    
    try:
        # Create batch processing indicator
        with open(batch_meta_path, "w") as f:
            json.dump({
                "batch_id": batch_id,
                "total_files": len(files),
                "status": "processing",
                "started_at": datetime.now().isoformat()
            }, f)
        
        # Upload all files with their metadata AND log to database one by one
        for file in files:
            try:
                # Save the file
                file_path = UPLOAD_DIR / file.filename
                with open(file_path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
                
                # Save metadata as sidecar JSON file
                meta = {
                    "departments": departments_list,
                    "countries": countries_list,
                    "uploaded_by": current_user.get("username"),
                    "batch_id": batch_id,
                    "upload_time": datetime.now().isoformat()
                }
                meta_path = file_path.with_suffix(file_path.suffix + ".meta.json")
                with open(meta_path, "w", encoding="utf-8") as f:
                    json.dump(meta, f, indent=2)
                
                # Log to database (one by one, just like /upload-file)
                conn = get_db()
                try:
                    with conn.cursor() as cursor:
                        # Get user ID
                        cursor.execute("SELECT user_id FROM users WHERE username = %s", (current_user.get("username"),))
                        user_result = cursor.fetchone()
                        user_id = user_result["user_id"] if user_result else 1
                        
                        # Insert into files table
                        cursor.execute("""
                            INSERT INTO files (file_name, file_type, uploaded_by, uploaded_by_username, department, access_level, file_path, 
                                             countries, departments, batch_id, upload_time)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            file.filename,
                            file_path.suffix.lstrip(".").lower(),
                            user_id,
                            current_user.get("username"),
                            ",".join(departments_list),
                            "FILTERED",
                            str(file_path),
                            ",".join(countries_list),
                            ",".join(departments_list),
                            batch_id,
                            datetime.now()
                        ))
                        
                        # INSERT INTO upload_file_logs table
                        cursor.execute("""
                            INSERT INTO upload_file_logs (file_name, file_type, uploaded_by, uploaded_by_username, department, access_level, file_path, countries, departments, batch_id, upload_time)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            file.filename,
                            file_path.suffix.lstrip(".").lower(),
                            user_id,
                            current_user.get("username"),
                            ",".join(departments_list),
                            "FILTERED",
                            str(file_path),
                            ",".join(countries_list),
                            ",".join(departments_list),
                            batch_id,
                            datetime.now()
                        ))
                        
                        conn.commit()
                finally:
                    conn.close()
                
                uploaded_files.append(file.filename)
                print(f"File {file.filename} uploaded and logged to both tables")
                
            except Exception as e:
                failed_files.append({"filename": file.filename, "error": str(e)})
                print(f"Failed to upload {file.filename}: {e}")
                
                # Clean up file if something goes wrong
                if 'file_path' in locals() and file_path.exists():
                    file_path.unlink()
                if 'meta_path' in locals() and meta_path.exists():
                    meta_path.unlink()
        
        # Update batch status to completed
        with open(batch_meta_path, "w") as f:
            json.dump({
                "batch_id": batch_id,
                "total_files": len(files),
                "uploaded_files": len(uploaded_files),
                "failed_files": len(failed_files),
                "status": "completed",
                "started_at": datetime.now().isoformat(),
                "completed_at": datetime.now().isoformat()
            }, f)
        
        # Rename to trigger watcher (remove .processing extension)
        final_batch_path = UPLOAD_DIR / f".batch_upload_{batch_id}.json"
        batch_meta_path.rename(final_batch_path)
        
        return {
            "message": f"Batch upload completed. {len(uploaded_files)} files uploaded successfully.",
            "uploaded_files": uploaded_files,
            "failed_files": failed_files,
            "batch_id": batch_id
        }
        
    except Exception as e:
        # Clean up on error
        if batch_meta_path.exists():
            batch_meta_path.unlink()
        raise HTTPException(status_code=500, detail=f"Batch upload failed: {str(e)}")


# Handl Logout
@app.post("/logout")
def logout(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        redis_client.delete(f"user_token:{user_id}")
        return {"message": "Logged out successfully"}
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


# Serve React frontend
# app.mount(
#     "/",
#     StaticFiles(directory=r"C:\Users\txcjs\OneDrive\Documents\Homework\Yr 3.1\ICP\Presentation\frontend\dist", html=True),
#     name="static"
# )

@app.post("/verify-admin-password")
async def verify_admin_password(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    try:
        # Get the user's stored password hash from database
        conn = get_db()
        with conn.cursor() as cursor:
            cursor.execute("SELECT password_hash FROM users WHERE user_id = %s", (current_user["user_id"],))
            user_record = cursor.fetchone()
            
            if not user_record:
                raise HTTPException(status_code=404, detail="User not found")
            # Hash the provided password and compare with stored hash
            if verify_password(data.get("password"), user_record["password_hash"]):
                return {"success": True}
            else:
                raise HTTPException(status_code=401, detail="Invalid password")
                
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Password verification failed: {str(e)}")
    finally:
        conn.close()

# for speech-to-textAdd commentMore actions
# change to medium or large for better performance only with GPU
model = whisper.load_model("base")

@app.post("/transcribe")
async def transcribe(file: UploadFile = File(...)):
    if not file:
        return {"text": "No file uploaded."}

    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix=".webm") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        # Load, process, and transcribe audio
        audio = whisper.load_audio(tmp_path)
        audio = whisper.pad_or_trim(audio)
        mel = whisper.log_mel_spectrogram(audio).to(model.device)

        # Detect language
        _, probs = model.detect_language(mel)
        detected_lang = max(probs, key=probs.get)
        print(f"Detected language: {detected_lang}")

        options = whisper.DecodingOptions(language=detected_lang)
        result = whisper.decode(model, mel, options)

        return {"text": result.text}
    finally:
        os.remove(tmp_path)

#chat search functionality
@app.get("/search")
async def search_chats(q: str = "", current_user: dict = Depends(get_current_user)):
    """
    Search chats by query, return grouped by unique conversation_id.
    Each result includes: conversation_id, preview (latest non-empty response), created_at (timestamp of that response), and the first non-empty title for that conversation.
    """
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            if not q.strip():
                # Recent chats (last 30 days)
                cursor.execute("""
                    SELECT conversation_id, MAX(created_at) as latest_time
                    FROM chatbot_logs
                    WHERE user_id = %s AND created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
                    GROUP BY conversation_id
                    ORDER BY latest_time DESC
                    LIMIT 30
                """, (current_user["user_id"],))
                rows = cursor.fetchall()
                results = []
                for row in rows:
                    # Get latest non-empty response as preview
                    cursor.execute("""
                        SELECT response, created_at FROM chatbot_logs
                        WHERE user_id = %s AND conversation_id = %s AND response IS NOT NULL AND response != ''
                        ORDER BY created_at DESC LIMIT 1
                    """, (current_user["user_id"], row["conversation_id"]))
                    preview_row = cursor.fetchone()
                    preview = preview_row["response"] if preview_row else ""
                    created_at = preview_row["created_at"] if preview_row else row["latest_time"]
                    # Always get the first non-empty title
                    cursor.execute("""
                        SELECT title FROM chatbot_logs
                        WHERE user_id = %s AND conversation_id = %s AND title IS NOT NULL AND title != ''
                        ORDER BY created_at ASC LIMIT 1
                    """, (current_user["user_id"], row["conversation_id"]))
                    title_row = cursor.fetchone()
                    title = title_row["title"] if title_row and title_row["title"] else "Untitled Chat"
                    results.append({
                        "conversation_id": str(row["conversation_id"]),
                        "preview": preview,
                        "created_at": created_at,
                        "title": title
                    })
                return results

            # If query, search logs and group by conversation_id
            cursor.execute("""
                SELECT conversation_id, MAX(created_at) as latest_time
                FROM chatbot_logs
                WHERE user_id = %s AND (query LIKE %s OR response LIKE %s)
                GROUP BY conversation_id
                ORDER BY latest_time DESC
                LIMIT 30
            """, (current_user["user_id"], f"%{q}%", f"%{q}%"))
            rows = cursor.fetchall()
            results = []
            for row in rows:
                # Get latest matching non-empty response as preview
                cursor.execute("""
                    SELECT response, created_at FROM chatbot_logs
                    WHERE user_id = %s AND conversation_id = %s AND response LIKE %s AND response IS NOT NULL AND response != ''
                    ORDER BY created_at DESC LIMIT 1
                """, (current_user["user_id"], row["conversation_id"], f"%{q}%"))
                preview_row = cursor.fetchone()
                preview = preview_row["response"] if preview_row else ""
                created_at = preview_row["created_at"] if preview_row else row["latest_time"]
                # Always get the first non-empty title
                cursor.execute("""
                    SELECT title FROM chatbot_logs
                    WHERE user_id = %s AND conversation_id = %s AND title IS NOT NULL AND title != ''
                    ORDER BY created_at ASC LIMIT 1
                """, (current_user["user_id"], row["conversation_id"]))
                title_row = cursor.fetchone()
                title = title_row["title"] if title_row and title_row["title"] else "Untitled Chat"
                results.append({
                    "conversation_id": str(row["conversation_id"]),
                    "preview": preview,
                    "created_at": created_at,
                    "title": title
                })
            return results
    finally:
        conn.close()

# for chat search functionality
@app.get("/chat-log/{log_id}")
def get_chat_log(log_id: int):
    connection = get_db()
    try:
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            sql = """
                SELECT query, response
                FROM chatbot_logs
                WHERE log_id = %s
            """
            cursor.execute(sql, (log_id,))
            rows = cursor.fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail="Log not found")

            messages = []
            for row in rows:
                messages.append({"role": "user", "content": row["query"]})
                messages.append({"role": "assistant", "content": row["response"]})

            return {"messages": messages}
    finally:
        connection.close()

# Serve the pipeline/misc folder as /static
app.mount(
    "/static",
    StaticFiles(directory=str(PROJECT_ROOT / "pipeline" / "misc")),
    name="static",
)

@app.put("/users/batch-update")
async def batch_update_users(
    request: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Batch update users. Only fields in 'data' will be updated for all user_ids.
    Permission checks are enforced per user.
    """
    user_ids = request.get("user_ids", [])
    data = request.get("data", {})

    conn = get_db()
    updated = []
    failed = []
    try:
        with conn.cursor() as cursor:
            for user_id in user_ids:
                cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
                target = cursor.fetchone()
                if not target:
                    failed.append({"user_id": user_id, "reason": "User not found"})
                    continue

                # Permission checks (same as single update)
                if current_user["role"] == "MANAGER":
                    if target["role"] == "ADMIN":
                        failed.append({"user_id": user_id, "reason": "Managers cannot edit admins"})
                        continue
                    if target["department"].strip().lower() != current_user["department"].strip().lower() or \
                       target["country"].strip().lower() != current_user["country"].strip().lower():
                        failed.append({"user_id": user_id, "reason": "Managers can only edit users in their own department and country"})
                        continue
                    if "role" in data and data["role"] == "ADMIN":
                        failed.append({"user_id": user_id, "reason": "Managers cannot set role to ADMIN"})
                        continue

                allowed_fields = ["email", "department", "country", "role"]
                updates = []
                params = []
                audit_logs = []
                now = datetime.now()

                for field in allowed_fields:
                    if field in data and data[field] != target.get(field):
                        updates.append(f"{field} = %s")
                        params.append(data[field])
                        audit_logs.append({
                            "target_username": target["username"],
                            "changed_by_username": current_user["username"],
                            "field_name": field,
                            "old_value": target.get(field),
                            "new_value": data[field],
                            "changed_at": now
                        })

                if not updates:
                    failed.append({"user_id": user_id, "reason": "No valid fields to update"})
                    continue

                updates.append("updated_at = %s")
                params.append(now)
                params.append(user_id)
                sql = f"UPDATE users SET {', '.join(updates)} WHERE user_id = %s"
                cursor.execute(sql, params)

                # Insert audit logs
                for log in audit_logs:
                    cursor.execute(
                        """
                        INSERT INTO user_update_logs (target_username, changed_by_username, field_name, old_value, new_value, changed_at)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (
                            log["target_username"],
                            log["changed_by_username"],
                            log["field_name"],
                            log["old_value"],
                            log["new_value"],
                            log["changed_at"]
                        )
                    )
                updated.append(user_id)
            conn.commit()
        return {"updated": updated, "failed": failed}
    finally:
        conn.close()
        
@app.put("/users/{user_id}")
async def update_user(
    user_id: int,
    data: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    # Fetch target user
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
            target = cursor.fetchone()
            if not target:
                raise HTTPException(status_code=404, detail="User not found")

            # Only ADMIN can edit anyone. MANAGER can only edit USER or MANAGER in their own dept/country, not ADMINs.
            if current_user["role"] == "MANAGER":
                if target["role"] == "ADMIN":
                    raise HTTPException(status_code=403, detail="Managers cannot edit admins")
                if target["department"].strip().lower() != current_user["department"].strip().lower() or \
                   target["country"].strip().lower() != current_user["country"].strip().lower():
                    raise HTTPException(status_code=403, detail="Managers can only edit users in their own department and country")
                # Managers cannot set anyone to ADMIN
                if "role" in data and data["role"] == "ADMIN":
                    raise HTTPException(status_code=403, detail="Managers cannot set role to ADMIN")

            # Update allowed fields
            allowed_fields = ["email", "department", "country", "role"]

            updates = []
            params = []
            audit_logs = []
            now = datetime.now()

            for field in allowed_fields:
                if field in data and data[field] != target.get(field):
                    updates.append(f"{field} = %s")
                    params.append(data[field])
                    # Prepare audit log for this field
                    audit_logs.append({
                        "target_username": target["username"],
                        "changed_by_username": current_user["username"],
                        "field_name": field,
                        "old_value": target.get(field),
                        "new_value": data[field],
                        "changed_at": now
                    })

            if not updates:
                raise HTTPException(status_code=400, detail="No valid fields to update")
            
            # Update the database
            updates.append("updated_at = %s")
            params.append(now)
            params.append(user_id)
            sql = f"UPDATE users SET {', '.join(updates)} WHERE user_id = %s"
            cursor.execute(sql, params)

            # Insert audit logs
            # Insert audit logs
            for log in audit_logs:
                cursor.execute(
                    """
                    INSERT INTO user_update_logs (target_username, changed_by_username, field_name, old_value, new_value, changed_at)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        log["target_username"],
                        log["changed_by_username"],
                        log["field_name"],
                        log["old_value"],
                        log["new_value"],
                        log["changed_at"]
                    )
                )
            conn.commit()
            return {"message": "User updated successfully"}
    finally:
        conn.close()

@app.delete("/users/{user_id}")
async def delete_user(
    user_id: int,
    current_user: dict = Depends(get_current_user)
):
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
            target = cursor.fetchone()
            if not target:
                raise HTTPException(status_code=404, detail="User not found")
            # Prevent users from deleting themselves
            if user_id == current_user["user_id"]:
                raise HTTPException(status_code=403, detail="You cannot delete your own account.")
            if current_user["role"] == "MANAGER":
                if target["role"] == "ADMIN":
                    raise HTTPException(status_code=403, detail="Managers cannot delete admins")
                # Managers can delete other managers and users in their own department and country
                if target["department"].strip().lower() != current_user["department"].strip().lower() or \
                   target["country"].strip().lower() != current_user["country"].strip().lower():
                    raise HTTPException(status_code=403, detail="Managers can only delete users in their own department and country")
            # Log deletion
            cursor.execute(
                """
                INSERT INTO user_deletion_logs (
                    deleted_username, deleted_email, department, role, country, deleted_by_username, deleted_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    target["username"],
                    target["email"],
                    target["department"],
                    target["role"],
                    target["country"],
                    current_user["username"],
                    datetime.now()
                )
            )
            cursor.execute("DELETE FROM users WHERE user_id = %s", (user_id,))
            conn.commit()
            return {"message": "User deleted successfully"}
    finally:
        conn.close()

@app.post("/users/batch-delete")
async def batch_delete_users(
    user_ids: List[int] = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """
    Batch delete users. Permission checks are enforced per user.
    """
    conn = get_db()
    deleted = []
    failed = []
    try:
        with conn.cursor() as cursor:
            for user_id in user_ids:
                cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
                target = cursor.fetchone()
                if not target:
                    failed.append({"user_id": user_id, "reason": "User not found"})
                    continue
                # Prevent users from deleting themselves
                if user_id == current_user["user_id"]:
                    failed.append({"user_id": user_id, "reason": "You cannot delete your own account."})
                    continue
                if current_user["role"] == "MANAGER":
                    if target["role"] == "ADMIN":
                        failed.append({"user_id": user_id, "reason": "Managers cannot delete admins"})
                        continue
                    if target["department"].strip().lower() != current_user["department"].strip().lower() or \
                       target["country"].strip().lower() != current_user["country"].strip().lower():
                        failed.append({"user_id": user_id, "reason": "Managers can only delete users in their own department and country"})
                        continue
                # Log deletion
                cursor.execute(
                    """
                    INSERT INTO user_deletion_logs (
                        deleted_username, deleted_email, department, role, country, deleted_by_username, deleted_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        target["username"],
                        target["email"],
                        target["department"],
                        target["role"],
                        target["country"],
                        current_user["username"],
                        datetime.now()
                    )
                )
                cursor.execute("DELETE FROM users WHERE user_id = %s", (user_id,))
                deleted.append(user_id)
            conn.commit()
        return {"deleted": deleted, "failed": failed}
    finally:
        conn.close()

        
@app.put("/update-file/{filename}")
async def update_file_metadata(
    filename: str,
    departments: List[str] = Form(...),
    countries: List[str] = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Updates file metadata (departments, countries) for a given file.
    - Updates all sidecar .meta.json files with the same base name
    - Updates the database record
    - Updates ChromaDB document metadata for this file
    - Logs changes to file_update_logs
    """
    file_path = UPLOAD_DIR / filename
    meta_path = file_path.with_suffix(file_path.suffix + ".meta.json")
    if not meta_path.exists():
        raise HTTPException(status_code=404, detail="Metadata file not found")

    # Load metadata for the file (for passing to update_docs_with_metadata)
    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)
    meta["departments"] = departments
    meta["countries"] = countries

    # Update DB and log changes
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            # Get file_id and current values for logging
            cursor.execute("SELECT file_id, departments, countries FROM files WHERE file_name = %s", (filename,))
            file_row = cursor.fetchone()
            if not file_row:
                raise HTTPException(status_code=404, detail="File not found in database")
            file_id = file_row["file_id"]
            old_departments = file_row["departments"] or ""
            old_countries = file_row["countries"] or ""

            # Update the files table
            cursor.execute(
                "UPDATE files SET departments=%s, countries=%s WHERE file_name=%s",
                (",".join(departments), ",".join(countries), filename)
            )

            now = datetime.now()
            # Log changes for departments
            if old_departments != ",".join(departments):
                cursor.execute(
                    """
                    INSERT INTO file_update_logs (file_id, file_name, changed_by_username, field_name, old_value, new_value, changed_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        file_id,
                        filename,
                        current_user.get("username"),
                        "departments",
                        old_departments,
                        ",".join(departments),
                        now
                    )
                )
            # Log changes for countries
            if old_countries != ",".join(countries):
                cursor.execute(
                    """
                    INSERT INTO file_update_logs (file_id, file_name, changed_by_username, field_name, old_value, new_value, changed_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        file_id,
                        filename,
                        current_user.get("username"),
                        "countries",
                        old_countries,
                        ",".join(countries),
                        now
                    )
                )
        conn.commit()
    finally:
        conn.close()

    # Update ChromaDB and ALL .meta.json files with the same base name
    update_docs_with_metadata(filename, meta)

    return {"message": "File metadata updated"}
    
def update_docs_with_metadata(filename, file_metadata):
    """
    Updates all Chroma documents whose 'title' in metadata matches the given filename (tries all common extensions).
    Sets specified department and country flags. Preserves 'uploaded_by' and 'upload_time' if already set.
    Also updates all .meta.json files with the same base name.
    """
    import os
    import glob
    from datetime import datetime

    ALL_DEPARTMENTS = [
        "Human Resource", "Admin & Operations", "Project Management", "Procurement",
        "IT", "Marketing", "Business Development", "Finance", "Service Delivery"
    ]

    ALL_COUNTRIES = [
        "Singapore", "United Kingdom", "United States", "Thailand", "Indonesia",
        "Korea", "China", "Japan", "Vietnam", "Myanmar"
    ]

    # Normalize the metadata format first
    normalized_metadata = normalize_metadata_format(file_metadata)
    
    # Use the normalized lists
    departments = normalized_metadata.get("departments", [])
    countries = normalized_metadata.get("countries", [])

    # Extract base filename without extension
    base_name = os.path.splitext(filename)[0]
    possible_exts = [".pdf", ".doc", ".docx", ".txt"]
    updated_any = False

    # Update ChromaDB docs for all possible extensions
    all_docs = chroma_collection.get(limit=1000)
    for ext in possible_exts:
        full_title = f"{base_name}{ext}"
        matching_indices = [
            i for i, metadata in enumerate(all_docs["metadatas"])
            if metadata.get("title", "").lower() == full_title.lower()
        ]
        if not matching_indices:
            print(f"No documents found with title '{full_title}'")
            continue

        for i in matching_indices:
            doc_id = all_docs["ids"][i]
            original_doc = all_docs["documents"][i]
            original_metadata = all_docs["metadatas"][i]
            updated_metadata = original_metadata.copy()

            # Update top-level metadata
            for dept in ALL_DEPARTMENTS:
                updated_metadata[dept] = "True" if dept in departments else "False"
            for country in ALL_COUNTRIES:
                updated_metadata[country] = "True" if country in countries else "False"
            if "uploaded_by" not in updated_metadata:
                updated_metadata["uploaded_by"] = file_metadata.get("uploaded_by", "")
            if "upload_time" not in updated_metadata:
                updated_metadata["upload_time"] = file_metadata.get("upload_time", datetime.now().isoformat())
            
            # Update all metadata locations
            if "_node_content" in original_metadata:
                try:
                    import json
                    node_content = json.loads(original_metadata["_node_content"])
                    
                    # Update the main embedded metadata
                    for dept in ALL_DEPARTMENTS:
                        node_content["metadata"][dept] = "True" if dept in departments else "False"
                    for country in ALL_COUNTRIES:
                        node_content["metadata"][country] = "True" if country in countries else "False"
                    
                    # Update relationships metadata
                    if "relationships" in node_content and "1" in node_content["relationships"]:
                        rel_metadata = node_content["relationships"]["1"].get("metadata", {})
                        for dept in ALL_DEPARTMENTS:
                            rel_metadata[dept] = "True" if dept in departments else "False"
                        for country in ALL_COUNTRIES:
                            rel_metadata[country] = "True" if country in countries else "False"
                        node_content["relationships"]["1"]["metadata"] = rel_metadata
                    
                    # Update the _node_content with all fixed metadata
                    updated_metadata["_node_content"] = json.dumps(node_content)
                    print(f"Fixed all metadata locations for {doc_id}")

                except Exception as e:
                    print(f"Error updating embedded metadata for {doc_id}: {e}")

            chroma_collection.update(
                ids=[doc_id],
                documents=[original_doc],
                metadatas=[updated_metadata]
            )
            print(f"Updated: {doc_id} ({original_metadata.get('title')})")
            updated_any = True

    if not updated_any:
        print("No matching documents found for any of the tried extensions.")

    # Update all .meta.json files with the same base name
    meta_files = glob.glob(str(UPLOAD_DIR / f"{base_name}.*.meta.json"))
    for meta_path in meta_files:
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
        meta["departments"] = departments
        meta["countries"] = countries
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2)

def delete_docs_by_filename(filename):
    """
    Deletes all Chroma documents whose 'title' in metadata matches the given filename 
    (tries all common extensions). Also deletes any associated .meta.json files.
    """
    import os
    import glob

    # Extract base filename without extension
    base_name = os.path.splitext(filename)[0]
    possible_exts = [".pdf", ".doc", ".docx", ".txt"]
    deleted_any = False

    # Check ChromaDB collection
    all_docs = chroma_collection.get(limit=1000)

    for ext in possible_exts:
        full_title = f"{base_name}{ext}"
        # Collect all matching doc IDs
        matching_ids = [
            all_docs["ids"][i]
            for i, metadata in enumerate(all_docs["metadatas"])
            if metadata.get("title", "").lower() == full_title.lower()
        ]

        if not matching_ids:
            print(f"No documents found with title '{full_title}'")
            continue

        # Delete all matching docs at once
        chroma_collection.delete(ids=matching_ids)
        print(f"Deleted documents with title '{full_title}' (IDs: {matching_ids})")
        deleted_any = True

    if deleted_any:
        print("Deletion successful.")
    else:
        print("No documents found for deletion.")

    # Delete associated .meta.json files
    meta_files = glob.glob(str(UPLOAD_DIR / f"{base_name}.*.meta.json"))
    for meta_path in meta_files:
        try:
            os.remove(meta_path)
            print(f"Deleted meta file: {meta_path}")
        except Exception as e:
            print(f"Error deleting meta file {meta_path}: {e}")

def normalize_metadata_format(metadata):
    """
    Normalize metadata format to ensure departments and countries are arrays of strings.
    Handles both old format (comma-separated strings) and new format (arrays).
    """
    if not isinstance(metadata, dict):
        return metadata
    
    normalized = metadata.copy()
    
    # Fix departments
    if "departments" in normalized:
        deps = normalized["departments"]
        if isinstance(deps, str):
            # Single string - split by comma
            normalized["departments"] = [dept.strip() for dept in deps.split(",") if dept.strip()]
        elif isinstance(deps, list):
            # List - check if it contains comma-separated strings
            new_departments = []
            for item in deps:
                if isinstance(item, str) and "," in item:
                    # Split comma-separated string
                    new_departments.extend([dept.strip() for dept in item.split(",") if dept.strip()])
                elif isinstance(item, str):
                    new_departments.append(item.strip())
            normalized["departments"] = new_departments
    
    # Fix countries
    if "countries" in normalized:
        countries = normalized["countries"]
        if isinstance(countries, str):
            # Single string - split by comma
            normalized["countries"] = [country.strip() for country in countries.split(",") if country.strip()]
        elif isinstance(countries, list):
            # List - check if it contains comma-separated strings
            new_countries = []
            for item in countries:
                if isinstance(item, str) and "," in item:
                    # Split comma-separated string
                    new_countries.extend([country.strip() for country in item.split(",") if country.strip()])
                elif isinstance(item, str):
                    new_countries.append(item.strip())
            normalized["countries"] = new_countries
    
    return normalized


# Password reset models
class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordResetConfirm(BaseModel):
    token: str
    user_id: int
    new_password: str


@app.post("/forgot-password")
async def forgot_password(request: PasswordResetRequest, req: Request):
    """Send password reset link to user's email"""
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            # Check if user exists
            cursor.execute("SELECT user_id, username FROM users WHERE email = %s", (request.email,))
            user = cursor.fetchone()
            
            # Log the password reset request attempt (regardless of whether user exists)
            user_ip = req.client.host
            user_agent = req.headers.get("user-agent", "")
            
            cursor.execute("""
                INSERT INTO password_reset_audit (
                    reset_type, target_email, target_username, user_ip, user_agent, reset_at
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                "REQUESTED",
                request.email,
                user["username"] if user else None,
                user_ip,
                user_agent,
                datetime.now()
            ))
            
            if not user:
                conn.commit()
                # Don't reveal whether email exists or not for security
                return {"message": "If the email exists, a password reset link has been sent."}
            
            # Generate reset token
            reset_token = secrets.token_urlsafe(32)
            reset_token_hash = hashlib.sha256(reset_token.encode()).hexdigest()
            
            # Store reset token in database (expires in 1 hour)
            expire_time = datetime.now() + timedelta(hours=1)
            cursor.execute("""
                INSERT INTO password_reset_tokens (user_id, username, token_hash, expires_at)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE token_hash = %s, expires_at = %s
            """, (user["user_id"], user["username"], reset_token_hash, expire_time, reset_token_hash, expire_time))

            conn.commit()
            
            # Create reset link
            reset_link = f"http://{REMOTE_IP}:3000/reset-password?token={reset_token}&user_id={user['user_id']}"
            
            # Email content
            subject = "Password Reset Request - Verztec HR System"
            body = f"""
            <html>
            <body>
                <h2>Password Reset Request</h2>
                <p>Hello {user['username']},</p>
                <p>You have requested to reset your password for the Verztec HR System.</p>
                <p>Click the link below to reset your password:</p>
                <p><a href="{reset_link}" style="background-color: #EAB308; color: black; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold;">Reset Password</a></p>
                <p>This link will expire in 1 hour.</p>
                <p>If you did not request this password reset, please ignore this email.</p>
                <br>
                <p>If you have any questions, please contact our IT support at <a href="mailto:it-support@verztec.com">it-support@verztec.com</a> or call +65 1234 5678.</p>
                <br>
                <p>Best regards,<br>Verztec HR System</p>
            </body>
            </html>
            """
            
            # Send email
            email_sent = send_email(request.email, subject, body, is_html=True)
            
            if email_sent:
                return {"message": "If the email exists, a password reset link has been sent."}
            else:
                raise HTTPException(status_code=500, detail="Failed to send email")
                
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in forgot_password: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        conn.close()

@app.post("/reset-password")
async def reset_password(request: PasswordResetConfirm, req: Request):
    """Reset password using token"""
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            # Verify token
            token_hash = hashlib.sha256(request.token.encode()).hexdigest()
            cursor.execute("""
                SELECT user_id FROM password_reset_tokens 
                WHERE user_id = %s AND token_hash = %s AND expires_at > %s
            """, (request.user_id, token_hash, datetime.now()))
            
            token_record = cursor.fetchone()
            if not token_record:
                raise HTTPException(status_code=400, detail="Invalid or expired reset token")
            
            # Get user details for audit log
            cursor.execute("SELECT username, email FROM users WHERE user_id = %s", (request.user_id,))
            user_details = cursor.fetchone()
            
            # Hash new password
            new_password_hash = hash_password(request.new_password)
            
            # Update password
            cursor.execute("""
                UPDATE users SET password_hash = %s, updated_at = %s 
                WHERE user_id = %s
            """, (new_password_hash, datetime.now(), request.user_id))
            
            # Log the successful password reset
            user_ip = req.client.host
            user_agent = req.headers.get("user-agent", "")
            
            cursor.execute("""
                INSERT INTO password_reset_audit (
                    reset_type, target_email, target_username, user_ip, user_agent, 
                    reset_token_used, reset_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                "COMPLETED",
                user_details["email"] if user_details else None,
                user_details["username"] if user_details else None,
                user_ip,
                user_agent,
                token_hash[:16] + "...",  # Store partial token for audit (first 16 chars)
                datetime.now()
            ))
            
            # Delete used token
            cursor.execute("DELETE FROM password_reset_tokens WHERE user_id = %s", (request.user_id,))
            
            conn.commit()
            
            return {"message": "Password reset successfully"}
            
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in reset_password: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        conn.close()

class ChatTitle(BaseModel):
    conversation_id: str
    title: str
    created_at: datetime

class ChatRenameRequest(BaseModel):
    title: str

class ChatLog(BaseModel):
    conversation_id: str
    content: str
    role: str
    created_at: datetime

@app.get("/chat-history/titles", response_model=List[ChatTitle])
async def get_chat_titles(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT conversation_id, MAX(created_at) as latest_time
                FROM chatbot_logs
                WHERE user_id = %s
                GROUP BY conversation_id
                ORDER BY latest_time DESC
            """, (current_user["user_id"],))
            rows = cursor.fetchall()
            titles = []
            for row in rows:
                # Get the title from the first message in the conversation
                cursor.execute("""
                    SELECT title FROM chatbot_logs
                    WHERE user_id = %s AND conversation_id = %s AND title IS NOT NULL
                    ORDER BY created_at ASC LIMIT 1
                """, (current_user["user_id"], row["conversation_id"]))
                title_row = cursor.fetchone()
                chat_title = title_row["title"] if title_row and title_row["title"] else f"Conversation {row['conversation_id']}"
                titles.append({
                    "conversation_id": str(row["conversation_id"]),
                    "title": chat_title,
                    "created_at": row["latest_time"]
                })
            return titles
    finally:
        conn.close()

@app.get("/chat-history/{conversation_id}", response_model=List[ChatLog])
async def get_chat_history(conversation_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT conversation_id, query, response, created_at
                FROM chatbot_logs
                WHERE user_id = %s AND conversation_id = %s
                ORDER BY created_at ASC
            """, (current_user["user_id"], conversation_id))
            rows = cursor.fetchall()
            if not rows:
                raise HTTPException(status_code=404, detail="Chat history not found")
            messages = []
            for row in rows:
                if row['query']:
                    messages.append({
                        "conversation_id": row['conversation_id'],
                        "content": row['query'],
                        "role": "user",
                        "created_at": row['created_at']
                    })
                if row['response']:
                    messages.append({
                        "conversation_id": row['conversation_id'],
                        "content": row['response'],
                        "role": "assistant", 
                        "created_at": row['created_at']
                    })
            return messages
    finally:
        conn.close()


# === Rename chat title ===
@app.put("/chat-history/{conversation_id}/rename")
async def rename_chat_title(
    conversation_id: str = FastAPIPath(...),
    req: ChatRenameRequest = Body(...),
    current_user: dict = Depends(get_current_user)
):
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "UPDATE chatbot_logs SET title = %s WHERE user_id = %s AND conversation_id = %s",
                (req.title, current_user["user_id"], conversation_id)
            )
            conn.commit()
        return {"message": "Chat title updated"}
    finally:
        conn.close()

# === Delete chat ===
@app.delete("/chat-history/{conversation_id}")
async def delete_chat(
    conversation_id: str = FastAPIPath(...),
    current_user: dict = Depends(get_current_user)
):
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "DELETE FROM chatbot_logs WHERE user_id = %s AND conversation_id = %s",
                (current_user["user_id"], conversation_id)
            )
            conn.commit()
        return {"message": "Chat deleted"}
    finally:
        conn.close()

class FeedbackCreate(BaseModel):
    category: str
    message: str
    rating: Optional[int] = None

@app.post("/feedback")
async def create_feedback(feedback: FeedbackCreate, current_user: dict = Depends(get_current_user)):
    try:
        connection = get_db()
        cursor = connection.cursor()
        
        # Include user_id from the authenticated user
        query = "INSERT INTO feedback (category, message, rating, user_id, created_at) VALUES (%s, %s, %s, %s, NOW())"
        cursor.execute(query, (feedback.category, feedback.message, feedback.rating, current_user["user_id"]))
        
        connection.commit()
        cursor.close()
        connection.close()
        return {"message": "Feedback submitted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class FeedbackStatusUpdate(BaseModel):
    status: str

class FeedbackResponse(BaseModel):
    id: int
    message: str
    category: str
    rating: Optional[int]
    status: str
    created_at: datetime
    updated_at: datetime

@app.get("/api/feedback", response_model=List[FeedbackResponse])
async def get_feedback(current_user: dict = Depends(get_current_user)):
    try:
        connection = get_db()
        cursor = connection.cursor()

        # Fetch all feedback from database
        query = """
        SELECT id, message, category, rating, status, created_at, updated_at 
        FROM feedback 
        ORDER BY created_at DESC
        """
        cursor.execute(query)
        feedbacks = cursor.fetchall()

        cursor.close()
        connection.close()

        return feedbacks
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching feedback: {str(e)}")

@app.put("/api/feedback/{feedback_id}/status")
async def update_feedback_status(
    feedback_id: int, 
    status_update: FeedbackStatusUpdate,
    current_user: dict = Depends(get_current_user)
):
    try:

        # Validate status
        valid_statuses = ["pending", "reviewed", "resolved"]
        if status_update.status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")

        connection = get_db()
        cursor = connection.cursor()

        # Check if feedback exists
        check_query = "SELECT id FROM feedback WHERE id = %s"
        cursor.execute(check_query, (feedback_id,))
        if not cursor.fetchone():
            cursor.close()
            connection.close()
            raise HTTPException(status_code=404, detail="Feedback not found")

        # Update feedback status
        update_query = "UPDATE feedback SET status = %s, updated_at = NOW() WHERE id = %s"
        cursor.execute(update_query, (status_update.status, feedback_id))
        connection.commit()
        cursor.close()
        connection.close()

        return {"message": "Feedback status updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating feedback status: {str(e)}")

class TranslationRequest(BaseModel):
    text: str
    from_lang: str
    to_lang: str

class TranslationResponse(BaseModel):
    translated_text: str
    from_lang: str
    to_lang: str

@app.post("/translate", response_model=TranslationResponse)
async def translate_text(request: TranslationRequest):
    try:
        translated = argostranslate.translate.translate(
            request.text, 
            request.from_lang, 
            request.to_lang
        )
        return TranslationResponse(
            translated_text=translated,
            from_lang=request.from_lang,
            to_lang=request.to_lang
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))