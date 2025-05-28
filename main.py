from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import ollama
from fastapi import FastAPI
from pydantic import BaseModel
from llama_index.llms.ollama import Ollama
from fastapi.middleware.cors import CORSMiddleware
from llama_index.embeddings.huggingface import HuggingFaceEmbedding
from llama_index.core import Settings
from llama_index.vector_stores.faiss import FaissVectorStore
import re
import faiss
from llama_index.core import VectorStoreIndex, StorageContext, load_index_from_storage
import os
from fastapi.responses import StreamingResponse
from fastapi import HTTPException, Depends
from fastapi.security import OAuth2PasswordBearer
from jose import jwt, JWTError
from enum import Enum
import bcrypt
import pymysql
import pymysql.cursors
from datetime import datetime, timedelta
from pydantic import EmailStr
from typing import Generator
import io
from fastapi import Request
from openpyxl import load_workbook


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Set up Ollama and HuggingFace embedding
Settings.embed_model = HuggingFaceEmbedding(model_name="intfloat/e5-large-v2")
embed_model = HuggingFaceEmbedding(model_name="intfloat/e5-large-v2")

# Instantiate the Ollama LLM
llm = Ollama(model="llama3.2:latest", request_timeout=120.0, temperature=0, context_window=4096)
Settings.llm = llm

# FAISS index loading
persist_dir = "../Embedded"
faiss_path = "faiss.index"
faiss_index = faiss.read_index(os.path.join(persist_dir, faiss_path))
vector_store = FaissVectorStore(faiss_index=faiss_index)
storage_context = StorageContext.from_defaults(persist_dir=persist_dir, vector_store=vector_store)
index = load_index_from_storage(storage_context)
query_engine = index.as_query_engine(similarity_top_k=5, streaming = True)

# This your llama-mini
light_llm = Ollama(
    model="llama3.2:1b",
    context_window=1024
)
# Prompt template
prompt_template = """
You are Verztec's AI HR assistant.

Your task is to extract all meaningful HR-related questions or concerns from the user's message.

For each item:
- Assign a concise label (e.g., "Leave entitlement", "Workplace harassment")
- Rewrite the question or concern in a short, clear form — **do not add explanations, notes, or assumptions**
- Only include relevant HR-related content, not exhaustive-(e.g., leave, claims, WFH, policies, misconduct, benefits, office or organisation-related matters)
- Always rewrite from the user's point of view using "I" instead of "you" unless specified
- Ignore all non-HR and off-topic content

⚠️ Your response must strictly follow this format:
1. **[Label]**: [Simplified question or concern]

❌ Do NOT include commentary, explanations, extra notes, or anything outside the format above.

If no HR-related questions or concerns are found, respond exactly: No HR-related questions or concerns detected.

Below is the raw user message, between <user></user> tags. Only process what's inside:

<user>
{user_input}
</user>
"""

# Pydantic model
class UserMessage(BaseModel):
    message: str

# Call LLM and parse output
def extract_questions(user_input: str) -> list[str]:
    prompt = prompt_template.format(user_input=user_input)
    response = llm.complete(prompt)
    raw_text = response.text.strip()

    # If no HR-related content detected
    if "no hr-related questions or concerns detected" in raw_text.lower():
        return []

    # Extract formatted entries
    question_list = []
    for line in raw_text.split("\n"):
        match = re.match(r"\d+\.\s+\**(.*?)\**:\s+(.*)", line.strip())
        if match:
            label, question = match.groups()
            question_list.append(f"{label.strip()}: {question.strip()}")

    return question_list

# Endpoint
@app.post("/process")
async def process_message(data: UserMessage):
    questions = extract_questions(data.message)

    response_text = ""
    if not questions:
        fallback_prompt = f"""
        The user said:

        "{data.message}"

        If it's not an HR question, feel free to respond naturally and politely — even if it's something casual or personal like "I love you".
        If it's unclear, gently suggest asking about HR topics like leave, benefits, claims, WFH, or company policies.
        Keep your reply friendly and human-like.
        """
        fallback = light_llm.complete(fallback_prompt).text.strip()
        response_text = fallback
        result = {"questions": [], "fallback": fallback}
    else:
        response_text = "\n".join(questions)
        result = {"questions": questions}

    # Log conversation here
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            sql = """
                INSERT INTO chatbot_logs (user_id, conversation_id, query, response)
                VALUES (%s, %s, %s, %s)
            """
            user_id = 1  # Replace with dynamic user id if available
            conversation_id = "1"  # Replace as needed

            cursor.execute(sql, (user_id, conversation_id, data.message, response_text))
        conn.commit()
    finally:
        conn.close()

    return result


@app.post("/stream")
async def stream_answer(data: UserMessage):
    user_prompt = data.message
    print(f"[STREAM] Querying with: {user_prompt}")

    response = query_engine.query(user_prompt)

    # This buffer will store the full response for logging
    full_response = []

    def stream_generator() -> Generator[str, None, None]:
        for token in response.response_gen:
            full_response.append(token)
            yield token

    # This wrapper allows us to perform logging after streaming is complete
    async def streaming_with_logging():
        generator = stream_generator()
        buffer = io.StringIO()
        for token in generator:
            buffer.write(token)
            yield token
        final_response = buffer.getvalue()

        # Log to DB after full response is available
        try:
            conn = get_db()
            with conn.cursor() as cursor:
                sql = """
                    INSERT INTO chatbot_logs (user_id, conversation_id, query, response)
                    VALUES (%s, %s, %s, %s)
                """
                user_id = 1
                conversation_id = "1"
                cursor.execute(sql, (user_id, conversation_id, user_prompt, final_response))
            conn.commit()
        finally:
            conn.close()

    return StreamingResponse(streaming_with_logging(), media_type="text/plain")


@app.post("/generate")
def generate(data: UserMessage):
    prompt = data.message
    response = ollama.chat(model="llama3.2", messages=[{"role": "user", "content": prompt}])
    return {"response": response["message"]["content"]}

# === Auth Configuration ===
SECRET_KEY = "verztec_secret"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

# === DB Connection ===
def get_db():
    return pymysql.connect(
        host="localhost",
        user="root",
        password="Tanhongkai123",
        database="verztec"
    )

# === Models ===
class Role(str, Enum):
    ADMIN = "ADMIN"
    MANAGER = "MANAGER"
    USER = "USER"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"

# === Utility ===
def verify_password(plain_pw, hashed_pw):
    return bcrypt.checkpw(plain_pw.encode('utf-8'), hashed_pw.encode('utf-8'))

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

# === Routes ===
@app.post("/login", response_model=TokenResponse)
def login_user(user: UserLogin, request: Request):
    db = get_db()
    cursor = db.cursor(pymysql.cursors.DictCursor)
    status = "FAILURE"
    user_id = None

    try:
        cursor.execute("SELECT * FROM users WHERE email = %s", (user.email,))
        user_record = cursor.fetchone()

        if user_record and verify_password(user.password, user_record["password_hash"]):
            user_id = user_record["user_id"]
            status = "SUCCESS"

            # Generate token
            token = create_access_token({
                "sub": user_record["username"],
                "role": user_record["role"]
            })

            # Log successful login
            cursor.execute("""
                INSERT INTO login_logs (user_id, status)
                VALUES (%s, %s)
            """, (user_id, status))

            db.commit()
            return {"access_token": token, "token_type": "bearer"}

        else:
            # Log failed login (if user exists or not)
            if user_record:
                user_id = user_record["user_id"]

            cursor.execute("""
                INSERT INTO login_logs (user_id, status)
                VALUES (%s, %s)
            """, (user_id, status))

            db.commit()
            raise HTTPException(status_code=401, detail="Invalid email or password")

    finally:
        cursor.close()
        db.close()

@app.get("/profile")
def read_profile(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if not username:
            raise HTTPException(status_code=401, detail="Invalid token")
        return {"message": f"Hello {username}, your token is valid!"}
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

class ChatLog(BaseModel):
    query: str
    response: str

# === Conversation Logging Endpoint ===
@app.post("/log_conversation/")
async def log_conversation(log: ChatLog):
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            sql = """
                INSERT INTO chatbot_logs (user_id, conversation_id, query, response)
                VALUES (%s, %s, %s, %s)
            """

            # Hardcoded user and conversation IDs
            user_id = 1
            conversation_id = "1"

            cursor.execute(sql, (
                user_id,
                conversation_id,
                log.query,
                log.response
            ))
        conn.commit()
        return {"message": "Conversation logged successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Logging failed: {str(e)}")
    finally:
        conn.close()
        
# User model
class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    department: str
    role: Role
    country: str
    
# Password hashing
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

# Upload XLSX and insert users
@app.post("/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        sheet = wb.active

        conn = get_db()
        cursor = conn.cursor()

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            username, email, password, department, role, country = row

            # Skip if email domain is not valid
            if not isinstance(email, str) or not email.endswith("@verztec.com"):
                continue

            hashed_pw = hash_password(password)

            try:
                cursor.execute("""
                    INSERT INTO users (username, email, password_hash, department, role, country)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (username, email, hashed_pw, department, role, country))
            except pymysql.err.IntegrityError:
                continue  # Skip duplicate entries

        conn.commit()
        cursor.close()
        conn.close()

        return {"message": "User data processed successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")