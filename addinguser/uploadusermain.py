from fastapi import FastAPI, UploadFile, File, HTTPException
from pydantic import BaseModel, EmailStr
from enum import Enum
from openpyxl import load_workbook
import bcrypt
import io
import mysql.connector

app = FastAPI()

#Database connection
def get_db():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="Tanhongkai123",
        database="verztec"
    )

class Role(str, Enum):
    ADMIN = "ADMIN"
    MANAGER = "MANAGER"
    USER = "USER"

class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    department: str
    role: Role
    country: str

#Password hashing for security
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

# Upload XLSX file and insert users to database
@app.post("/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        sheet = wb.active

        db = get_db()
        cursor = db.cursor()

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            username, email, password, department, role, country = row

             # Skipping invalid emails
            if not email.endswith("@verztec.com"):
                continue

            hashed_pw = hash_password(password)

            try:
                cursor.execute("""
                    INSERT INTO users (username, email, password_hash, department, role, country)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (username, email, hashed_pw, department, role, country))
            except mysql.connector.IntegrityError:
                continue

        db.commit()
        cursor.close()
        db.close()

        return {"message": "User data processed successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
